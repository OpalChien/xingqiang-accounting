from __future__ import annotations

import calendar
import io
import json
import os
import re
import sqlite3
import uuid
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from urllib.error import URLError
from urllib.request import Request, urlopen
from zoneinfo import ZoneInfo

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components


APP_TZ = ZoneInfo("Asia/Taipei")
BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
COMPANY_NAME = "興強科技有限公司"
COMPANY_SHORT_NAME = "興強科技"
COMPANY_TAX_ID = "54155450"
COMPANY_PHONE = "02-27031206"
COMPANY_ADDRESS = "台北市大安區和平東路一段242號2F"
COMPANY_WEBSITE = "https://0227031206.tw66.com.tw"
COMPANY_BUSINESS_LINES = [
    "德國 Kerb-Konus Vertriebs 緊固件代理",
    "日本內橋 Uchihashi ELCUT 溫度保險絲代理",
    "電子器材設備批發與零組件銷售",
]

TRADE_FLOWS = ["出口", "進口"]
ACCOUNT_SIDES = ["應收", "應付"]
SETTLEMENT_CYCLES = ["當下結", "月結", "雙月結", "半年結"]
CURRENCIES = ["TWD", "USD", "EUR", "JPY", "CNY", "HKD", "GBP", "AUD", "CAD", "SGD"]
ERP_CUSTOMER_REQUIRED_COLUMNS = {"Customer ID", "English name", "Credit days", "Currency"}

TABLE_COLUMNS = [
    "id",
    "trade_flow",
    "account_side",
    "customer_id",
    "counterparty",
    "invoice_no",
    "order_no",
    "shipment_no",
    "item_description",
    "currency",
    "amount_original",
    "exchange_rate",
    "amount_twd",
    "settlement_cycle",
    "invoice_date",
    "grace_days",
    "due_date",
    "paid_amount_twd",
    "payment_date",
    "bank_account",
    "owner",
    "notes",
    "created_at",
    "updated_at",
]

DISPLAY_COLUMNS = {
    "id": "ID",
    "trade_flow": "進出口",
    "account_side": "應收/應付",
    "customer_id": "客戶編號",
    "counterparty": "客戶/供應商",
    "invoice_no": "發票號碼",
    "order_no": "訂單號碼",
    "shipment_no": "提單/報關號碼",
    "item_description": "品名/摘要",
    "currency": "幣別",
    "amount_original": "原幣金額",
    "exchange_rate": "匯率",
    "amount_twd": "台幣金額",
    "settlement_cycle": "結帳方式",
    "invoice_date": "交易日期",
    "grace_days": "付款天數",
    "due_date": "到期日",
    "paid_amount_twd": "已收/已付金額",
    "payment_date": "收付款日期",
    "bank_account": "銀行/帳戶",
    "owner": "承辦人",
    "notes": "備註",
    "created_at": "建立時間",
    "updated_at": "更新時間",
    "outstanding_twd": "未結金額",
    "computed_status": "狀態",
    "days_overdue": "逾期天數",
    "settlement_period": "結帳期間",
    "aging_bucket": "帳齡",
}

IMPORT_ALIASES = {
    "ID": "id",
    "id": "id",
    "進出口": "trade_flow",
    "trade_flow": "trade_flow",
    "應收/應付": "account_side",
    "account_side": "account_side",
    "客戶編號": "customer_id",
    "Customer ID": "customer_id",
    "customer_id": "customer_id",
    "客戶/供應商": "counterparty",
    "客戶": "counterparty",
    "供應商": "counterparty",
    "counterparty": "counterparty",
    "English name": "counterparty",
    "發票號碼": "invoice_no",
    "發票": "invoice_no",
    "invoice_no": "invoice_no",
    "訂單號碼": "order_no",
    "訂單": "order_no",
    "order_no": "order_no",
    "提單/報關號碼": "shipment_no",
    "提單": "shipment_no",
    "報關號碼": "shipment_no",
    "shipment_no": "shipment_no",
    "品名/摘要": "item_description",
    "品名": "item_description",
    "摘要": "item_description",
    "item_description": "item_description",
    "幣別": "currency",
    "currency": "currency",
    "原幣金額": "amount_original",
    "金額": "amount_original",
    "amount_original": "amount_original",
    "匯率": "exchange_rate",
    "exchange_rate": "exchange_rate",
    "台幣金額": "amount_twd",
    "amount_twd": "amount_twd",
    "結帳方式": "settlement_cycle",
    "結帳條件": "settlement_cycle",
    "settlement_cycle": "settlement_cycle",
    "交易日期": "invoice_date",
    "發票日期": "invoice_date",
    "invoice_date": "invoice_date",
    "付款天數": "grace_days",
    "grace_days": "grace_days",
    "到期日": "due_date",
    "due_date": "due_date",
    "已收/已付金額": "paid_amount_twd",
    "已收金額": "paid_amount_twd",
    "已付金額": "paid_amount_twd",
    "paid_amount_twd": "paid_amount_twd",
    "收付款日期": "payment_date",
    "付款日期": "payment_date",
    "payment_date": "payment_date",
    "銀行/帳戶": "bank_account",
    "銀行帳戶": "bank_account",
    "bank_account": "bank_account",
    "承辦人": "owner",
    "owner": "owner",
    "備註": "notes",
    "notes": "notes",
}

CUSTOMER_ALIASES = {
    "Customer ID": "customer_id",
    "客戶編號": "customer_id",
    "customer_id": "customer_id",
    "English name": "english_name",
    "英文名稱": "english_name",
    "客戶英文名稱": "english_name",
    "english_name": "english_name",
    "Chinese name": "chinese_name",
    "中文名稱": "chinese_name",
    "chinese_name": "chinese_name",
    "Currency": "currency",
    "幣別": "currency",
    "currency": "currency",
    "Credit days": "credit_days",
    "ERP信用天數": "credit_days",
    "信用天數": "credit_days",
    "credit_days": "credit_days",
    "付款天數": "grace_days",
    "grace_days": "grace_days",
    "結帳方式": "settlement_cycle",
    "settlement_cycle": "settlement_cycle",
    "Payment Terms": "payment_terms",
    "付款條件": "payment_terms",
    "payment_terms": "payment_terms",
    "SalesPerson": "sales_person",
    "業務": "sales_person",
    "sales_person": "sales_person",
    "Bus. Type": "business_type",
    "客戶類別": "business_type",
    "business_type": "business_type",
    "Shipment Terms": "shipment_terms",
    "出貨條件": "shipment_terms",
    "shipment_terms": "shipment_terms",
    "Contact Person": "contact_person",
    "聯絡人": "contact_person",
    "contact_person": "contact_person",
    "Phone No.": "phone",
    "電話": "phone",
    "phone": "phone",
    "Email": "email",
    "email": "email",
}


def main() -> None:
    st.set_page_config(
        page_title=f"{COMPANY_SHORT_NAME}記帳平台",
        page_icon="帳",
        layout="wide",
        initial_sidebar_state="expanded",
    )
    inject_styles()
    init_db()

    st.title(f"{COMPANY_SHORT_NAME}記帳平台")
    st.caption("進出口代理、應收應付、結帳週期與 Excel 對帳一次整理。")
    render_company_strip()
    init_session_flags()

    if not st.session_state["data_loaded"]:
        render_initial_load_screen()
        return

    raw_df = fetch_transactions()
    render_download_reminder(raw_df)
    enriched_df = enrich_transactions(raw_df)
    filtered_df = render_sidebar_filters(enriched_df)

    tabs = st.tabs(["總覽", "新增帳款", "已收已付", "明細查詢", "Excel 匯入匯出", "設定"])
    with tabs[0]:
        render_dashboard(filtered_df)
    with tabs[1]:
        render_entry_and_payment(raw_df)
    with tabs[2]:
        render_payment_records(raw_df)
    with tabs[3]:
        render_simple_detail_table(filtered_df)
    with tabs[4]:
        render_excel_tools(raw_df)
    with tabs[5]:
        render_settings()


def init_session_flags() -> None:
    st.session_state.setdefault("data_loaded", False)
    st.session_state.setdefault("data_downloaded", False)
    st.session_state.setdefault("data_changed", False)


def mark_data_loaded() -> None:
    st.session_state["data_loaded"] = True
    st.session_state["data_downloaded"] = False
    st.session_state["data_changed"] = False


def mark_data_changed() -> None:
    st.session_state["data_downloaded"] = False
    st.session_state["data_changed"] = True


def mark_data_downloaded() -> None:
    st.session_state["data_downloaded"] = True
    st.session_state["data_changed"] = False


def report_file_name() -> str:
    return f"{COMPANY_SHORT_NAME}_對帳備份_{today().isoformat()}.xlsx"


def render_initial_load_screen() -> None:
    st.warning("每次開始記帳前，請先上傳上一次下載的 Excel 對帳備份；也可先載入 ERP 客戶主檔。")
    st.caption("交易明細是每天對帳主檔；ERP 客戶主檔會用來自動帶幣別與付款天數。")

    col1, col2 = st.columns([1.2, 1])
    with col1:
        uploaded = st.file_uploader("載入 Excel 對帳備份或 ERP 客戶主檔", type=["xlsx", "xls"], key="initial_excel_upload")
        if uploaded is not None:
            try:
                payload = parse_workbook_payload(uploaded)
                records = payload["transactions"]
                customers = payload["customers"]
                if records:
                    st.markdown("**交易明細預覽**")
                    preview = enrich_transactions(pd.DataFrame(records))
                    st.dataframe(preview.head(10), hide_index=True, use_container_width=True)
                if customers:
                    st.markdown("**ERP 客戶主檔預覽**")
                    st.dataframe(format_customer_preview(pd.DataFrame(customers).head(10)), hide_index=True, use_container_width=True)
                if st.button("確認載入資料", type="primary", use_container_width=True):
                    if records:
                        replace_transactions(records)
                    else:
                        replace_transactions([])
                    replace_customers(customers)
                    mark_data_loaded()
                    st.success(f"已載入 {len(records)} 筆交易、{len(customers)} 筆 ERP 客戶資料。")
                    st.rerun()
            except Exception as exc:
                st.error(f"載入失敗：{exc}")

    with col2:
        st.download_button(
            "下載空白匯入範本",
            data=build_template_workbook(),
            file_name=f"{COMPANY_SHORT_NAME}_匯入範本.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        local_df = fetch_transactions()
        local_customers_df = fetch_customers()
        st.button(
            "載入本機暫存資料",
            disabled=local_df.empty and local_customers_df.empty,
            use_container_width=True,
            on_click=mark_data_loaded,
            help="本機測試時可用；正式部署建議每次上傳 Excel 備份。",
        )
        if st.button("從空白開始", use_container_width=True):
            replace_transactions([])
            replace_customers([])
            mark_data_loaded()
            st.rerun()


def render_download_reminder(raw_df: pd.DataFrame) -> None:
    should_warn = bool(st.session_state.get("data_loaded")) and not bool(st.session_state.get("data_downloaded"))
    inject_close_warning(should_warn)

    if raw_df.empty:
        st.info("目前沒有交易明細。若已載入 ERP 客戶主檔，仍建議下載今日 Excel 備份。")

    if should_warn:
        st.warning("離開或關閉視窗前，請先下載今日 Excel 對帳備份。")
    else:
        st.success("今日 Excel 備份已下載；如果後續又新增或登記收付款，系統會再次提醒。")

    st.download_button(
        f"下載今日 Excel：{report_file_name()}",
        data=build_report_workbook(enrich_transactions(raw_df)),
        file_name=report_file_name(),
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        on_click=mark_data_downloaded,
    )


def inject_close_warning(should_warn: bool) -> None:
    if not should_warn:
        components.html(
            """
            <script>
            try {
              const target = window.parent || window;
              if (target.__xingqiangBeforeUnloadHandler) {
                target.removeEventListener("beforeunload", target.__xingqiangBeforeUnloadHandler);
                target.__xingqiangBeforeUnloadHandler = null;
              }
            } catch (error) {}
            </script>
            """,
            height=0,
        )
        return

    components.html(
        """
        <script>
        try {
          const target = window.parent || window;
          if (target.__xingqiangBeforeUnloadHandler) {
            target.removeEventListener("beforeunload", target.__xingqiangBeforeUnloadHandler);
          }
          target.__xingqiangBeforeUnloadHandler = function(event) {
            event.preventDefault();
            event.returnValue = "離開前請確認已下載今天的 Excel 對帳備份。";
            return event.returnValue;
          };
          target.addEventListener("beforeunload", target.__xingqiangBeforeUnloadHandler);
        } catch (error) {}
        </script>
        """,
        height=0,
    )


def render_company_strip() -> None:
    with st.container(border=True):
        col1, col2, col3 = st.columns([1.25, 1, 1])
        with col1:
            st.markdown(f"**{COMPANY_NAME}**")
            st.markdown(
                f"<span class='small-muted'>{COMPANY_BUSINESS_LINES[0]}；{COMPANY_BUSINESS_LINES[1]}</span>",
                unsafe_allow_html=True,
            )
        with col2:
            st.markdown(f"統編：`{COMPANY_TAX_ID}`")
            st.markdown(f"電話：`{COMPANY_PHONE}`")
        with col3:
            st.link_button("公司網站", COMPANY_WEBSITE, use_container_width=True)
            st.markdown(
                f"<span class='small-muted'>{COMPANY_WEBSITE}</span><br>"
                f"<span class='small-muted'>{COMPANY_ADDRESS}</span>",
                unsafe_allow_html=True,
            )


def inject_styles() -> None:
    st.markdown(
        """
        <style>
        .stApp {
            background: #f7f8f5;
            color: #1e2528;
        }
        h1, h2, h3 {
            letter-spacing: 0;
        }
        [data-testid="stMetric"] {
            background: #ffffff;
            border: 1px solid #dde3dd;
            border-radius: 8px;
            padding: 14px 16px;
            box-shadow: 0 1px 2px rgba(31, 38, 35, 0.04);
        }
        [data-testid="stMetricValue"] {
            font-size: 1.35rem;
        }
        div[data-testid="stVerticalBlockBorderWrapper"] {
            border-radius: 8px;
        }
        .status-pill {
            display: inline-block;
            padding: 2px 8px;
            border-radius: 999px;
            border: 1px solid #cad2cb;
            background: #fff;
            font-size: 12px;
        }
        .small-muted {
            color: #62706b;
            font-size: 13px;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def init_db() -> None:
    DATA_DIR.mkdir(exist_ok=True)
    with get_connection() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS transactions (
                id TEXT PRIMARY KEY,
                trade_flow TEXT NOT NULL,
                account_side TEXT NOT NULL,
                customer_id TEXT,
                counterparty TEXT NOT NULL,
                invoice_no TEXT,
                order_no TEXT,
                shipment_no TEXT,
                item_description TEXT,
                currency TEXT NOT NULL,
                amount_original REAL NOT NULL,
                exchange_rate REAL NOT NULL,
                amount_twd REAL NOT NULL,
                settlement_cycle TEXT NOT NULL,
                invoice_date TEXT NOT NULL,
                grace_days INTEGER NOT NULL,
                due_date TEXT NOT NULL,
                paid_amount_twd REAL NOT NULL,
                payment_date TEXT,
                bank_account TEXT,
                owner TEXT,
                notes TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        ensure_column(conn, "transactions", "customer_id", "TEXT")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS customers (
                customer_id TEXT PRIMARY KEY,
                english_name TEXT NOT NULL,
                chinese_name TEXT,
                currency TEXT NOT NULL,
                credit_days INTEGER NOT NULL,
                settlement_cycle TEXT NOT NULL,
                grace_days INTEGER NOT NULL,
                payment_terms TEXT,
                sales_person TEXT,
                business_type TEXT,
                shipment_terms TEXT,
                contact_person TEXT,
                phone TEXT,
                email TEXT,
                imported_at TEXT NOT NULL
            )
            """
        )


def ensure_column(conn: sqlite3.Connection, table: str, column: str, definition: str) -> None:
    existing = {row["name"] for row in conn.execute(f"PRAGMA table_info({table})").fetchall()}
    if column not in existing:
        conn.execute(f"ALTER TABLE {table} ADD COLUMN {column} {definition}")


def get_connection() -> sqlite3.Connection:
    db_path = Path(os.getenv("ACCOUNTING_DB_PATH", str(DATA_DIR / "accounting.db")))
    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn


def fetch_transactions() -> pd.DataFrame:
    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT *
            FROM transactions
            ORDER BY date(invoice_date) DESC, datetime(created_at) DESC
            """
        ).fetchall()

    if not rows:
        return pd.DataFrame(columns=TABLE_COLUMNS)

    df = pd.DataFrame([dict(row) for row in rows])
    for col in TABLE_COLUMNS:
        if col not in df.columns:
            df[col] = None
    return df[TABLE_COLUMNS]


def fetch_customers() -> pd.DataFrame:
    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT *
            FROM customers
            ORDER BY english_name, customer_id
            """
        ).fetchall()
    if not rows:
        return pd.DataFrame(
            columns=[
                "customer_id",
                "english_name",
                "chinese_name",
                "currency",
                "credit_days",
                "settlement_cycle",
                "grace_days",
                "payment_terms",
                "sales_person",
                "business_type",
                "shipment_terms",
                "contact_person",
                "phone",
                "email",
                "imported_at",
            ]
        )
    return pd.DataFrame([dict(row) for row in rows])


def build_customer_options(customers_df: pd.DataFrame) -> dict[str, str]:
    options: dict[str, str] = {"手動輸入": ""}
    if customers_df.empty:
        return options

    for _, row in customers_df.sort_values(["english_name", "customer_id"]).iterrows():
        customer_id = clean_text(row.get("customer_id"))
        english_name = clean_text(row.get("english_name"))
        if not customer_id or not english_name:
            continue
        currency = normalize_currency(row.get("currency")) or "TWD"
        cycle = normalize_cycle(row.get("settlement_cycle") or "月結")
        days = int(safe_float(row.get("grace_days")))
        options[f"{english_name} ({customer_id}) | {currency} | {cycle}+{days}天"] = customer_id
    return options


def get_customer_record(customers_df: pd.DataFrame, customer_id: str) -> dict[str, Any] | None:
    if customers_df.empty or not customer_id:
        return None
    matches = customers_df[customers_df["customer_id"] == customer_id]
    if matches.empty:
        return None
    return matches.iloc[0].to_dict()


def replace_customers(records: list[dict[str, Any]]) -> None:
    with get_connection() as conn:
        conn.execute("DELETE FROM customers")
    upsert_customers(records)


def upsert_customers(records: list[dict[str, Any]]) -> None:
    now = datetime.now(APP_TZ).isoformat(timespec="seconds")
    normalized_records = []
    for record in records:
        customer_id = clean_text(record.get("customer_id"))
        english_name = clean_text(record.get("english_name"))
        if not customer_id or not english_name:
            continue
        credit_days = max(int(safe_float(record.get("credit_days"))), 0)
        grace_days = int(safe_float(record.get("grace_days"), default=credit_days))
        normalized_records.append(
            {
                "customer_id": customer_id,
                "english_name": english_name,
                "chinese_name": clean_text(record.get("chinese_name")),
                "currency": normalize_currency(record.get("currency")),
                "credit_days": credit_days,
                "settlement_cycle": normalize_cycle(record.get("settlement_cycle") or "月結"),
                "grace_days": max(grace_days, 0),
                "payment_terms": clean_text(record.get("payment_terms")),
                "sales_person": clean_text(record.get("sales_person")),
                "business_type": clean_text(record.get("business_type")),
                "shipment_terms": clean_text(record.get("shipment_terms")),
                "contact_person": clean_text(record.get("contact_person")),
                "phone": clean_text(record.get("phone")),
                "email": clean_text(record.get("email")),
                "imported_at": now,
            }
        )
    if not normalized_records:
        return
    with get_connection() as conn:
        conn.executemany(
            """
            INSERT INTO customers (
                customer_id, english_name, chinese_name, currency, credit_days,
                settlement_cycle, grace_days, payment_terms, sales_person, business_type,
                shipment_terms, contact_person, phone, email, imported_at
            )
            VALUES (
                :customer_id, :english_name, :chinese_name, :currency, :credit_days,
                :settlement_cycle, :grace_days, :payment_terms, :sales_person, :business_type,
                :shipment_terms, :contact_person, :phone, :email, :imported_at
            )
            ON CONFLICT(customer_id) DO UPDATE SET
                english_name = excluded.english_name,
                chinese_name = excluded.chinese_name,
                currency = excluded.currency,
                credit_days = excluded.credit_days,
                settlement_cycle = excluded.settlement_cycle,
                grace_days = excluded.grace_days,
                payment_terms = excluded.payment_terms,
                sales_person = excluded.sales_person,
                business_type = excluded.business_type,
                shipment_terms = excluded.shipment_terms,
                contact_person = excluded.contact_person,
                phone = excluded.phone,
                email = excluded.email,
                imported_at = excluded.imported_at
            """,
            normalized_records,
        )


def upsert_transaction(record: dict[str, Any]) -> str:
    now = datetime.now(APP_TZ).isoformat(timespec="seconds")
    record_id = str(record.get("id") or uuid.uuid4())
    invoice_date = coerce_date(record.get("invoice_date")) or today()
    grace_days = int(record.get("grace_days") or 0)
    amount_original = safe_float(record.get("amount_original"))
    exchange_rate = safe_float(record.get("exchange_rate"), default=1.0)
    if exchange_rate <= 0:
        exchange_rate, _ = get_live_exchange_rate(record.get("currency", "TWD"))
    amount_twd = round(amount_original * exchange_rate, 2)
    due_date = calculate_due_date(invoice_date, record.get("settlement_cycle", "月結"), grace_days)
    payment_date = coerce_date(record.get("payment_date"))

    normalized = {
        "id": record_id,
        "trade_flow": normalize_choice(record.get("trade_flow"), TRADE_FLOWS, "出口"),
        "account_side": normalize_choice(record.get("account_side"), ACCOUNT_SIDES, "應收"),
        "customer_id": clean_text(record.get("customer_id")),
        "counterparty": clean_text(record.get("counterparty")) or "未命名",
        "invoice_no": clean_text(record.get("invoice_no")),
        "order_no": clean_text(record.get("order_no")),
        "shipment_no": clean_text(record.get("shipment_no")),
        "item_description": clean_text(record.get("item_description")),
        "currency": normalize_currency(record.get("currency")),
        "amount_original": amount_original,
        "exchange_rate": exchange_rate,
        "amount_twd": amount_twd,
        "settlement_cycle": normalize_cycle(record.get("settlement_cycle")),
        "invoice_date": invoice_date.isoformat(),
        "grace_days": grace_days,
        "due_date": due_date.isoformat(),
        "paid_amount_twd": safe_float(record.get("paid_amount_twd")),
        "payment_date": payment_date.isoformat() if payment_date else None,
        "bank_account": clean_text(record.get("bank_account")),
        "owner": clean_text(record.get("owner")),
        "notes": clean_text(record.get("notes")),
        "updated_at": now,
    }

    with get_connection() as conn:
        exists = conn.execute("SELECT 1 FROM transactions WHERE id = ?", (record_id,)).fetchone()
        normalized["created_at"] = (
            conn.execute("SELECT created_at FROM transactions WHERE id = ?", (record_id,)).fetchone()["created_at"]
            if exists
            else now
        )
        conn.execute(
            """
            INSERT INTO transactions (
                id, trade_flow, account_side, customer_id, counterparty, invoice_no, order_no, shipment_no,
                item_description, currency, amount_original, exchange_rate, amount_twd,
                settlement_cycle, invoice_date, grace_days, due_date, paid_amount_twd,
                payment_date, bank_account, owner, notes, created_at, updated_at
            )
            VALUES (
                :id, :trade_flow, :account_side, :customer_id, :counterparty, :invoice_no, :order_no, :shipment_no,
                :item_description, :currency, :amount_original, :exchange_rate, :amount_twd,
                :settlement_cycle, :invoice_date, :grace_days, :due_date, :paid_amount_twd,
                :payment_date, :bank_account, :owner, :notes, :created_at, :updated_at
            )
            ON CONFLICT(id) DO UPDATE SET
                trade_flow = excluded.trade_flow,
                account_side = excluded.account_side,
                customer_id = excluded.customer_id,
                counterparty = excluded.counterparty,
                invoice_no = excluded.invoice_no,
                order_no = excluded.order_no,
                shipment_no = excluded.shipment_no,
                item_description = excluded.item_description,
                currency = excluded.currency,
                amount_original = excluded.amount_original,
                exchange_rate = excluded.exchange_rate,
                amount_twd = excluded.amount_twd,
                settlement_cycle = excluded.settlement_cycle,
                invoice_date = excluded.invoice_date,
                grace_days = excluded.grace_days,
                due_date = excluded.due_date,
                paid_amount_twd = excluded.paid_amount_twd,
                payment_date = excluded.payment_date,
                bank_account = excluded.bank_account,
                owner = excluded.owner,
                notes = excluded.notes,
                updated_at = excluded.updated_at
            """,
            normalized,
        )
    return record_id


def delete_transaction(record_id: str) -> None:
    with get_connection() as conn:
        conn.execute("DELETE FROM transactions WHERE id = ?", (record_id,))


def register_payment(record_id: str, payment_amount_twd: float, payment_date_value: date) -> None:
    payment_amount = max(safe_float(payment_amount_twd), 0)
    if payment_amount <= 0:
        raise ValueError("請填寫本次收/付款金額。")

    with get_connection() as conn:
        row = conn.execute("SELECT * FROM transactions WHERE id = ?", (record_id,)).fetchone()
        if row is None:
            raise ValueError("找不到這筆帳款。")

    record = dict(row)
    total_amount = safe_float(record.get("amount_twd"))
    current_paid = safe_float(record.get("paid_amount_twd"))
    record["paid_amount_twd"] = min(total_amount, current_paid + payment_amount)
    record["payment_date"] = payment_date_value
    upsert_transaction(record)


@st.cache_data(ttl=60 * 60 * 6, show_spinner=False)
def get_live_exchange_rate(currency: str) -> tuple[float, str]:
    currency_code = normalize_currency(currency)
    if currency_code == "TWD":
        return 1.0, "TWD"

    url = f"https://open.er-api.com/v6/latest/{currency_code}"
    try:
        request = Request(url, headers={"User-Agent": "XingqiangAccounting/1.0"})
        with urlopen(request, timeout=8) as response:
            payload = json.loads(response.read().decode("utf-8"))
        if payload.get("result") != "success":
            return 1.0, "匯率服務暫時無法使用"
        rate = safe_float(payload.get("rates", {}).get("TWD"), default=0)
        if rate <= 0:
            return 1.0, "找不到 TWD 匯率"
        updated = clean_text(payload.get("time_last_update_utc")) or "latest"
        return rate, updated
    except (OSError, URLError, TimeoutError, json.JSONDecodeError):
        return 1.0, "匯率服務暫時無法連線"


def replace_transactions(records: list[dict[str, Any]]) -> None:
    with get_connection() as conn:
        conn.execute("DELETE FROM transactions")
    for record in records:
        upsert_transaction(record)


def append_transactions(records: list[dict[str, Any]]) -> None:
    for record in records:
        upsert_transaction(record)


def calculate_due_date(invoice_date: date, settlement_cycle: str, grace_days: int = 0) -> date:
    cycle = normalize_cycle(settlement_cycle)
    if cycle == "當下結":
        base = invoice_date
    elif cycle == "月結":
        base = month_end(invoice_date.year, invoice_date.month)
    elif cycle == "雙月結":
        end_month = ((invoice_date.month - 1) // 2 + 1) * 2
        base = month_end(invoice_date.year, end_month)
    elif cycle == "半年結":
        base = date(invoice_date.year, 6, 30) if invoice_date.month <= 6 else date(invoice_date.year, 12, 31)
    else:
        base = month_end(invoice_date.year, invoice_date.month)
    return base + timedelta(days=max(int(grace_days), 0))


def month_end(year: int, month: int) -> date:
    return date(year, month, calendar.monthrange(year, month)[1])


def settlement_period(invoice_date_value: Any, settlement_cycle: str) -> str:
    invoice_date = coerce_date(invoice_date_value)
    if not invoice_date:
        return ""
    cycle = normalize_cycle(settlement_cycle)
    if cycle == "當下結":
        return invoice_date.isoformat()
    if cycle == "月結":
        return f"{invoice_date.year}-{invoice_date.month:02d}"
    if cycle == "雙月結":
        start_month = ((invoice_date.month - 1) // 2) * 2 + 1
        return f"{invoice_date.year}-{start_month:02d}/{start_month + 1:02d}"
    half = "H1" if invoice_date.month <= 6 else "H2"
    return f"{invoice_date.year}-{half}"


def enrich_transactions(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        empty = pd.DataFrame(columns=list(DISPLAY_COLUMNS.values()))
        return empty

    enriched = df.copy()
    for col in TABLE_COLUMNS:
        if col not in enriched.columns:
            enriched[col] = None

    enriched["invoice_date"] = enriched["invoice_date"].apply(lambda value: coerce_date(value) or today())
    enriched["settlement_cycle"] = enriched["settlement_cycle"].apply(normalize_cycle)
    enriched["grace_days"] = pd.to_numeric(enriched["grace_days"], errors="coerce").fillna(0).astype(int)
    enriched["exchange_rate"] = pd.to_numeric(enriched["exchange_rate"], errors="coerce").fillna(1)
    enriched["amount_original"] = pd.to_numeric(enriched["amount_original"], errors="coerce").fillna(0)
    amount_twd = pd.to_numeric(enriched["amount_twd"], errors="coerce")
    enriched["amount_twd"] = amount_twd.fillna(enriched["amount_original"] * enriched["exchange_rate"])
    enriched["due_date"] = enriched.apply(
        lambda row: coerce_date(row.get("due_date"))
        or calculate_due_date(row["invoice_date"], row["settlement_cycle"], int(row["grace_days"])),
        axis=1,
    )

    for date_col in ["invoice_date", "due_date", "payment_date", "created_at", "updated_at"]:
        if date_col in enriched.columns:
            enriched[date_col] = pd.to_datetime(enriched[date_col], errors="coerce")

    numeric_cols = ["amount_original", "exchange_rate", "amount_twd", "paid_amount_twd", "grace_days"]
    for col in numeric_cols:
        enriched[col] = pd.to_numeric(enriched[col], errors="coerce").fillna(0)

    enriched["outstanding_twd"] = (enriched["amount_twd"] - enriched["paid_amount_twd"]).clip(lower=0).round(2)
    local_today = pd.Timestamp(today())
    due_dates = pd.to_datetime(enriched["due_date"], errors="coerce")
    overdue_days = (local_today - due_dates).dt.days.fillna(0).astype(int)
    enriched["days_overdue"] = overdue_days.where(enriched["outstanding_twd"] > 0, 0).clip(lower=0)
    enriched["computed_status"] = enriched.apply(compute_status, axis=1)
    enriched["settlement_period"] = enriched.apply(
        lambda row: settlement_period(row.get("invoice_date"), row.get("settlement_cycle")),
        axis=1,
    )
    enriched["aging_bucket"] = enriched["days_overdue"].apply(aging_bucket)

    display = enriched.rename(columns=DISPLAY_COLUMNS)
    ordered = [label for label in DISPLAY_COLUMNS.values() if label in display.columns]
    return display[ordered]


def compute_status(row: pd.Series) -> str:
    outstanding = safe_float(row.get("outstanding_twd"))
    paid = safe_float(row.get("paid_amount_twd"))
    days_overdue = int(row.get("days_overdue") or 0)
    if outstanding <= 0:
        return "已結清"
    if days_overdue > 0 and paid > 0:
        return "逾期部分"
    if days_overdue > 0:
        return "逾期"
    if paid > 0:
        return "部分"
    return "未結"


def aging_bucket(days_overdue: int) -> str:
    if days_overdue <= 0:
        return "未到期"
    if days_overdue <= 30:
        return "逾期 1-30"
    if days_overdue <= 60:
        return "逾期 31-60"
    if days_overdue <= 90:
        return "逾期 61-90"
    return "逾期 90+"


def render_sidebar_filters(df: pd.DataFrame) -> pd.DataFrame:
    st.sidebar.header("篩選")
    if df.empty:
        st.sidebar.info("目前沒有資料。")
        return df

    min_date = pd.to_datetime(df["交易日期"], errors="coerce").min()
    max_date = pd.to_datetime(df["交易日期"], errors="coerce").max()
    default_start = min_date.date() if pd.notna(min_date) else today().replace(day=1)
    default_end = max_date.date() if pd.notna(max_date) else today()

    date_range = st.sidebar.date_input("交易日期", value=(default_start, default_end))
    selected_flow = st.sidebar.multiselect("進出口", TRADE_FLOWS, default=TRADE_FLOWS)
    selected_side = st.sidebar.multiselect("應收/應付", ACCOUNT_SIDES, default=ACCOUNT_SIDES)
    selected_cycle = st.sidebar.multiselect("結帳方式", SETTLEMENT_CYCLES, default=SETTLEMENT_CYCLES)
    status_options = sorted(df["狀態"].dropna().unique().tolist())
    selected_status = st.sidebar.multiselect("狀態", status_options, default=status_options)
    counterparty_query = st.sidebar.text_input("客戶/供應商搜尋")

    filtered = df.copy()
    if isinstance(date_range, tuple) and len(date_range) == 2:
        start_date, end_date = date_range
        tx_dates = pd.to_datetime(filtered["交易日期"], errors="coerce").dt.date
        filtered = filtered[(tx_dates >= start_date) & (tx_dates <= end_date)]
    if selected_flow:
        filtered = filtered[filtered["進出口"].isin(selected_flow)]
    if selected_side:
        filtered = filtered[filtered["應收/應付"].isin(selected_side)]
    if selected_cycle:
        filtered = filtered[filtered["結帳方式"].isin(selected_cycle)]
    if selected_status:
        filtered = filtered[filtered["狀態"].isin(selected_status)]
    if counterparty_query:
        filtered = filtered[
            filtered["客戶/供應商"].astype(str).str.contains(counterparty_query, case=False, na=False)
        ]
    return filtered


def render_dashboard(df: pd.DataFrame) -> None:
    if df.empty:
        st.info("尚未有帳款資料，可以先到「新增帳款」或「Excel 匯入匯出」建立資料。")
        return

    receivable = df[df["應收/應付"] == "應收"]["未結金額"].sum()
    payable = df[df["應收/應付"] == "應付"]["未結金額"].sum()
    overdue = df[df["逾期天數"] > 0]["未結金額"].sum()
    settled = df[df["狀態"] == "已結清"]["台幣金額"].sum()

    kpi1, kpi2, kpi3, kpi4 = st.columns(4)
    kpi1.metric("應收未結", money(receivable))
    kpi2.metric("應付未結", money(payable))
    kpi3.metric("逾期未結", money(overdue))
    kpi4.metric("已結清金額", money(settled))

    left, right = st.columns([1.25, 1])
    with left:
        st.subheader("到期月份")
        monthly = df.copy()
        monthly["到期月份"] = pd.to_datetime(monthly["到期日"], errors="coerce").dt.strftime("%Y-%m")
        monthly_pivot = (
            monthly.pivot_table(
                index="到期月份",
                columns="應收/應付",
                values="未結金額",
                aggfunc="sum",
                fill_value=0,
            )
            .sort_index()
            .tail(12)
        )
        st.bar_chart(monthly_pivot, height=280)

    with right:
        st.subheader("結帳方式")
        cycle_summary = (
            df.pivot_table(
                index="結帳方式",
                columns="應收/應付",
                values="未結金額",
                aggfunc="sum",
                fill_value=0,
            )
            .reindex(SETTLEMENT_CYCLES)
            .fillna(0)
        )
        st.dataframe(format_money_columns(cycle_summary.reset_index()), hide_index=True, use_container_width=True)

    st.subheader("優先對帳")
    priority = df[df["未結金額"] > 0].copy()
    priority = priority.sort_values(["逾期天數", "到期日"], ascending=[False, True]).head(12)
    show_cols = ["狀態", "應收/應付", "客戶/供應商", "發票號碼", "結帳方式", "到期日", "未結金額", "逾期天數"]
    st.dataframe(
        format_money_columns(priority[show_cols]),
        hide_index=True,
        use_container_width=True,
    )


def render_entry_and_payment(raw_df: pd.DataFrame) -> None:
    add_tab, edit_tab = st.tabs(["新增帳款", "編輯與收付款"])
    with add_tab:
        saved = render_simple_transaction_form(None, "add_record")
        if saved:
            st.success("已新增帳款。")
            st.rerun()

    with edit_tab:
        if raw_df.empty:
            st.info("目前沒有可編輯的帳款。")
            return
        options = build_record_options(raw_df)
        selected_label = st.selectbox("選擇帳款", list(options.keys()))
        selected_id = options[selected_label]
        selected_row = raw_df[raw_df["id"] == selected_id].iloc[0].to_dict()
        saved = render_simple_transaction_form(selected_row, "edit_record")
        if saved:
            st.success("已更新帳款。")
            st.rerun()

        st.divider()
        confirm_delete = st.checkbox("確認刪除此筆帳款")
        if st.button("刪除選取帳款", disabled=not confirm_delete, type="secondary"):
            delete_transaction(selected_id)
            mark_data_changed()
            st.success("已刪除。")
            st.rerun()


def render_transaction_form(existing: dict[str, Any] | None, form_key: str) -> bool:
    existing = existing or {}
    with st.form(form_key, clear_on_submit=existing == {}):
        col1, col2, col3 = st.columns(3)
        with col1:
            trade_flow = st.selectbox(
                "進出口",
                TRADE_FLOWS,
                index=index_of(TRADE_FLOWS, existing.get("trade_flow"), 0),
            )
            default_side = "應收" if trade_flow == "出口" else "應付"
            account_side = st.selectbox(
                "應收/應付",
                ACCOUNT_SIDES,
                index=index_of(ACCOUNT_SIDES, existing.get("account_side"), ACCOUNT_SIDES.index(default_side)),
            )
            counterparty = st.text_input("客戶/供應商", value=existing.get("counterparty", ""))
            invoice_no = st.text_input("發票號碼", value=existing.get("invoice_no", ""))
        with col2:
            order_no = st.text_input("訂單號碼", value=existing.get("order_no", ""))
            shipment_no = st.text_input("提單/報關號碼", value=existing.get("shipment_no", ""))
            item_description = st.text_input("品名/摘要", value=existing.get("item_description", ""))
            owner = st.text_input("承辦人", value=existing.get("owner", ""))
        with col3:
            currency_default = clean_text(existing.get("currency")).upper() or "TWD"
            currency_options = CURRENCIES if currency_default in CURRENCIES else [currency_default] + CURRENCIES
            currency = st.selectbox("幣別", currency_options, index=0 if currency_default not in CURRENCIES else CURRENCIES.index(currency_default))
            amount_original = st.number_input(
                "原幣金額",
                min_value=0.0,
                value=float(existing.get("amount_original") or 0),
                step=1000.0,
                format="%.2f",
            )
            exchange_rate = st.number_input(
                "匯率",
                min_value=0.0,
                value=float(existing.get("exchange_rate") or 1),
                step=0.01,
                format="%.6f",
            )
            st.metric("台幣金額", money(amount_original * exchange_rate))

        col4, col5, col6 = st.columns(3)
        with col4:
            invoice_date = st.date_input("交易日期", value=coerce_date(existing.get("invoice_date")) or today())
            settlement_cycle = st.selectbox(
                "結帳方式",
                SETTLEMENT_CYCLES,
                index=index_of(SETTLEMENT_CYCLES, existing.get("settlement_cycle"), 1),
            )
        with col5:
            grace_days = st.number_input(
                "付款天數",
                min_value=0,
                value=int(existing.get("grace_days") or 0),
                step=1,
            )
            projected_due = calculate_due_date(invoice_date, settlement_cycle, int(grace_days))
            st.metric("到期日", projected_due.isoformat())
        with col6:
            paid_amount_twd = st.number_input(
                "已收/已付金額",
                min_value=0.0,
                value=float(existing.get("paid_amount_twd") or 0),
                step=1000.0,
                format="%.2f",
            )
            has_payment_date = st.checkbox("填寫收付款日期", value=bool(existing.get("payment_date")))
            payment_date = None
            if has_payment_date:
                payment_date = st.date_input("收付款日期", value=coerce_date(existing.get("payment_date")) or today())

        bank_account = st.text_input("銀行/帳戶", value=existing.get("bank_account", ""))
        notes = st.text_area("備註", value=existing.get("notes", ""), height=84)

        submitted = st.form_submit_button("儲存帳款", type="primary")
        if not submitted:
            return False
        if not clean_text(counterparty):
            st.error("請填寫客戶/供應商。")
            return False
        upsert_transaction(
            {
                "id": existing.get("id"),
                "trade_flow": trade_flow,
                "account_side": account_side,
                "counterparty": counterparty,
                "invoice_no": invoice_no,
                "order_no": order_no,
                "shipment_no": shipment_no,
                "item_description": item_description,
                "currency": currency,
                "amount_original": amount_original,
                "exchange_rate": exchange_rate,
                "settlement_cycle": settlement_cycle,
                "invoice_date": invoice_date,
                "grace_days": int(grace_days),
                "paid_amount_twd": paid_amount_twd,
                "payment_date": payment_date,
                "bank_account": bank_account,
                "owner": owner,
                "notes": notes,
            }
        )
        mark_data_changed()
        return True


def render_simple_transaction_form(existing: dict[str, Any] | None, form_key: str) -> bool:
    existing = existing or {}
    record_id = clean_text(existing.get("id")) or "new"
    prefix = f"{form_key}_{record_id}"

    st.markdown("#### 簡化記帳")
    st.caption("先填對帳一定會用到的資料；訂單、提單、承辦人等細節可放到進階欄位。")

    customers_df = fetch_customers()
    customer_id_default = clean_text(existing.get("customer_id"))
    selected_customer: dict[str, Any] | None = None
    selected_customer_id = customer_id_default
    if not customers_df.empty:
        customer_options = build_customer_options(customers_df)
        option_labels = list(customer_options.keys())
        default_label = next(
            (label for label, customer_id in customer_options.items() if customer_id == customer_id_default),
            "手動輸入",
        )
        selected_label = st.selectbox(
            "套用 ERP 客戶資料",
            option_labels,
            index=option_labels.index(default_label) if default_label in option_labels else 0,
            key=f"{prefix}_customer_picker",
        )
        selected_customer_id = customer_options[selected_label]
        selected_customer = get_customer_record(customers_df, selected_customer_id)
        if selected_customer:
            st.caption(
                "ERP 客戶條件："
                f"{selected_customer_id} | "
                f"{normalize_currency(selected_customer.get('currency'))} | "
                f"{normalize_cycle(selected_customer.get('settlement_cycle'))}+"
                f"{int(safe_float(selected_customer.get('grace_days')))}天"
            )

    counterparty_default = (
        clean_text(selected_customer.get("english_name")) if selected_customer else clean_text(existing.get("counterparty"))
    )
    currency_default = normalize_currency(
        selected_customer.get("currency") if selected_customer else (clean_text(existing.get("currency")) or "TWD")
    )
    settlement_default = (
        normalize_cycle(selected_customer.get("settlement_cycle")) if selected_customer else normalize_cycle(existing.get("settlement_cycle"))
    )
    grace_days_default = int(
        safe_float(
            selected_customer.get("grace_days") if selected_customer else existing.get("grace_days"),
            default=0,
        )
    )
    owner_default = clean_text(selected_customer.get("sales_person")) if selected_customer else clean_text(existing.get("owner"))
    customer_key_part = selected_customer_id or "manual"

    basic_col, amount_col, settlement_col = st.columns(3)
    with basic_col:
        invoice_date = st.date_input(
            "日期",
            value=coerce_date(existing.get("invoice_date")) or today(),
            key=f"{prefix}_invoice_date",
        )
        trade_flow = st.selectbox(
            "進出口",
            TRADE_FLOWS,
            index=index_of(TRADE_FLOWS, existing.get("trade_flow"), 0),
            key=f"{prefix}_trade_flow",
        )
        default_side = "應收" if trade_flow == "出口" else "應付"
        account_side = st.selectbox(
            "帳款類型",
            ACCOUNT_SIDES,
            index=index_of(ACCOUNT_SIDES, existing.get("account_side"), ACCOUNT_SIDES.index(default_side)),
            key=f"{prefix}_account_side",
        )
        customer_id_value = st.text_input(
            "客戶編號",
            value=selected_customer_id,
            key=f"{prefix}_{customer_key_part}_customer_id",
            disabled=selected_customer is not None,
        )
        counterparty = st.text_input(
            "客戶/供應商",
            value=counterparty_default,
            key=f"{prefix}_{customer_key_part}_counterparty",
        )
        invoice_no = st.text_input(
            "發票號碼",
            value=clean_text(existing.get("invoice_no")),
            key=f"{prefix}_invoice_no",
        )

    with amount_col:
        currency_options = CURRENCIES if currency_default in CURRENCIES else [currency_default] + CURRENCIES
        currency = st.selectbox(
            "幣別",
            currency_options,
            index=0 if currency_default not in CURRENCIES else CURRENCIES.index(currency_default),
            key=f"{prefix}_{customer_key_part}_currency",
        )
        amount_original = st.number_input(
            "原幣金額",
            min_value=0.0,
            value=float(existing.get("amount_original") or 0),
            step=1000.0,
            format="%.2f",
            key=f"{prefix}_amount_original",
        )
        live_rate, rate_note = get_live_exchange_rate(currency)
        if currency == "TWD":
            exchange_rate = 1.0
            st.caption("台幣不用換算。")
        else:
            has_existing_rate = bool(existing.get("id")) and safe_float(existing.get("exchange_rate")) > 0
            rate_mode = st.radio(
                "匯率",
                ["自動抓匯率", "手動輸入"],
                index=1 if has_existing_rate else 0,
                horizontal=True,
                key=f"{prefix}_rate_mode",
            )
            if rate_mode == "自動抓匯率":
                exchange_rate = live_rate
                if exchange_rate == 1.0 and "latest" not in rate_note:
                    st.warning(rate_note)
                else:
                    st.caption(f"線上匯率：1 {currency} = {exchange_rate:,.4f} TWD")
            else:
                exchange_rate = st.number_input(
                    "手動匯率",
                    min_value=0.0,
                    value=float(existing.get("exchange_rate") or live_rate or 1),
                    step=0.01,
                    format="%.6f",
                    key=f"{prefix}_exchange_rate",
                )
        amount_twd = round(amount_original * exchange_rate, 2)
        st.metric("自動換算台幣", money(amount_twd))

    with settlement_col:
        settlement_cycle = st.selectbox(
            "結帳方式",
            SETTLEMENT_CYCLES,
            index=index_of(SETTLEMENT_CYCLES, settlement_default, 1),
            key=f"{prefix}_{customer_key_part}_settlement_cycle",
        )
        grace_days = st.number_input(
            "付款天數",
            min_value=0,
            value=max(grace_days_default, 0),
            step=1,
            key=f"{prefix}_{customer_key_part}_grace_days",
        )
        projected_due = calculate_due_date(invoice_date, settlement_cycle, int(grace_days))
        st.metric("到期日", projected_due.isoformat())
        paid_amount_twd = st.number_input(
            "已收/已付金額",
            min_value=0.0,
            value=float(existing.get("paid_amount_twd") or 0),
            step=1000.0,
            format="%.2f",
            key=f"{prefix}_paid_amount_twd",
        )
        st.metric("未結金額", money(max(amount_twd - paid_amount_twd, 0)))

    with st.expander("進階欄位", expanded=False):
        adv1, adv2 = st.columns(2)
        with adv1:
            order_no = st.text_input("訂單號碼", value=clean_text(existing.get("order_no")), key=f"{prefix}_order_no")
            shipment_no = st.text_input(
                "提單/報關號碼",
                value=clean_text(existing.get("shipment_no")),
                key=f"{prefix}_shipment_no",
            )
            item_description = st.text_input(
                "品名/摘要",
                value=clean_text(existing.get("item_description")),
                key=f"{prefix}_item_description",
            )
            owner = st.text_input("承辦人", value=owner_default, key=f"{prefix}_{customer_key_part}_owner")
        with adv2:
            bank_account = st.text_input(
                "銀行/帳戶",
                value=clean_text(existing.get("bank_account")),
                key=f"{prefix}_bank_account",
            )
            has_payment_date = st.checkbox(
                "填寫收付款日期",
                value=bool(existing.get("payment_date")),
                key=f"{prefix}_has_payment_date",
            )
            payment_date = None
            if has_payment_date:
                payment_date = st.date_input(
                    "收付款日期",
                    value=coerce_date(existing.get("payment_date")) or today(),
                    key=f"{prefix}_payment_date",
                )
            notes = st.text_area("備註", value=clean_text(existing.get("notes")), height=84, key=f"{prefix}_notes")

    if st.button("儲存帳款", type="primary", key=f"{prefix}_save"):
        if not clean_text(counterparty):
            st.error("請填寫客戶/供應商。")
            return False
        if amount_original <= 0:
            st.error("請填寫金額。")
            return False
        upsert_transaction(
            {
                "id": existing.get("id"),
                "trade_flow": trade_flow,
                "account_side": account_side,
                "customer_id": customer_id_value,
                "counterparty": counterparty,
                "invoice_no": invoice_no,
                "order_no": order_no,
                "shipment_no": shipment_no,
                "item_description": item_description,
                "currency": currency,
                "amount_original": amount_original,
                "exchange_rate": exchange_rate,
                "settlement_cycle": settlement_cycle,
                "invoice_date": invoice_date,
                "grace_days": int(grace_days),
                "paid_amount_twd": paid_amount_twd,
                "payment_date": payment_date,
                "bank_account": bank_account,
                "owner": owner,
                "notes": notes,
            }
        )
        mark_data_changed()
        return True
    return False


def render_simple_detail_table(df: pd.DataFrame) -> None:
    if df.empty:
        st.info("沒有符合條件的明細。")
        return

    st.subheader("簡化明細")
    visible_cols = [
        "狀態",
        "進出口",
        "應收/應付",
        "客戶編號",
        "客戶/供應商",
        "發票號碼",
        "幣別",
        "原幣金額",
        "匯率",
        "台幣金額",
        "結帳方式",
        "交易日期",
        "到期日",
        "收付款日期",
        "已收/已付金額",
        "未結金額",
    ]
    view = df[[col for col in visible_cols if col in df.columns]].copy()
    st.dataframe(format_money_columns(view), hide_index=True, use_container_width=True, height=420)

    with st.expander("完整欄位", expanded=False):
        st.dataframe(format_money_columns(df.copy()), hide_index=True, use_container_width=True, height=520)

    st.subheader("客戶/供應商彙總")
    summary = (
        df.pivot_table(
            index=["客戶/供應商", "應收/應付"],
            values=["台幣金額", "已收/已付金額", "未結金額"],
            aggfunc="sum",
            fill_value=0,
        )
        .reset_index()
        .sort_values("未結金額", ascending=False)
    )
    st.dataframe(format_money_columns(summary), hide_index=True, use_container_width=True)


def render_payment_records(raw_df: pd.DataFrame) -> None:
    if raw_df.empty:
        st.info("目前沒有帳款資料。")
        return

    df = enrich_transactions(raw_df)
    register_tab, history_tab = st.tabs(["登記收/付款", "已收已付清單"])

    with register_tab:
        outstanding = df[df["未結金額"] > 0].copy()
        if outstanding.empty:
            st.success("目前沒有未結帳款。")
        else:
            st.subheader("登記收款或付款")
            st.caption("選一筆未結帳款，填收付款日期與本次金額。日期會寫進對帳報表。")
            options = build_payment_options(outstanding)
            selected_label = st.selectbox("選擇未結帳款", list(options.keys()))
            selected_id = options[selected_label]
            selected_display = df[df["ID"] == selected_id].iloc[0]
            outstanding_amount = safe_float(selected_display["未結金額"])
            action_name = "收款" if selected_display["應收/應付"] == "應收" else "付款"

            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("帳款類型", selected_display["應收/應付"])
                st.metric("未結金額", money(outstanding_amount))
            with col2:
                customer_label = clean_text(selected_display.get("客戶編號"))
                if customer_label:
                    customer_label = f"{customer_label} / {selected_display['客戶/供應商']}"
                else:
                    customer_label = selected_display["客戶/供應商"]
                st.metric("客戶/供應商", customer_label)
                st.metric("到期日", selected_display["到期日"].date().isoformat() if pd.notna(selected_display["到期日"]) else "")
            with col3:
                st.metric("結帳方式", selected_display["結帳方式"])
                st.metric("狀態", selected_display["狀態"])

            pay_col, date_col = st.columns(2)
            with date_col:
                payment_date = st.date_input(f"{action_name}日期", value=today(), key="payment_record_date")
            with pay_col:
                payment_amount = st.number_input(
                    f"本次{action_name}金額",
                    min_value=0.0,
                    value=float(outstanding_amount),
                    step=1000.0,
                    format="%.2f",
                    key="payment_record_amount",
                )

            if st.button(f"登記{action_name}", type="primary", use_container_width=True):
                try:
                    register_payment(selected_id, payment_amount, payment_date)
                    mark_data_changed()
                    st.success(f"已登記{action_name}，日期：{payment_date.isoformat()}。")
                    st.rerun()
                except ValueError as exc:
                    st.error(str(exc))

    with history_tab:
        paid = df[df["已收/已付金額"] > 0].copy()
        if paid.empty:
            st.info("目前還沒有已收或已付紀錄。")
            return
        paid = paid.sort_values(["收付款日期", "交易日期"], ascending=[False, False], na_position="last")
        visible_cols = [
            "狀態",
            "應收/應付",
            "客戶編號",
            "客戶/供應商",
            "發票號碼",
            "收付款日期",
            "已收/已付金額",
            "未結金額",
            "台幣金額",
            "幣別",
            "原幣金額",
            "結帳方式",
            "交易日期",
            "備註",
        ]
        st.dataframe(
            format_money_columns(paid[[col for col in visible_cols if col in paid.columns]]),
            hide_index=True,
            use_container_width=True,
            height=520,
        )


def build_payment_options(df: pd.DataFrame) -> dict[str, str]:
    options: dict[str, str] = {}
    for _, row in df.sort_values(["到期日", "客戶/供應商"]).iterrows():
        due = row["到期日"].date().isoformat() if pd.notna(row["到期日"]) else ""
        invoice_no = clean_text(row.get("發票號碼")) or "無發票"
        customer_id = clean_text(row.get("客戶編號"))
        party = f"{customer_id} {row['客戶/供應商']}" if customer_id else row["客戶/供應商"]
        label = (
            f"{due} | {row['應收/應付']} | {party} | "
            f"{invoice_no} | 未結 {money(row['未結金額'])}"
        )
        options[label] = row["ID"]
    return options


def build_record_options(df: pd.DataFrame) -> dict[str, str]:
    options: dict[str, str] = {}
    enriched = enrich_transactions(df)
    for _, row in enriched.iterrows():
        customer_id = clean_text(row.get("客戶編號"))
        party = f"{customer_id} {row['客戶/供應商']}" if customer_id else row["客戶/供應商"]
        label = (
            f"{row['交易日期'].date() if pd.notna(row['交易日期']) else ''} | "
            f"{party} | {row['發票號碼'] or '無發票'} | "
            f"{row['應收/應付']} {money(row['未結金額'])}"
        )
        options[label] = row["ID"]
    return options


def render_detail_table(df: pd.DataFrame) -> None:
    if df.empty:
        st.info("沒有符合條件的明細。")
        return

    st.subheader("交易明細")
    visible_cols = [
        "狀態",
        "進出口",
        "應收/應付",
        "客戶編號",
        "客戶/供應商",
        "發票號碼",
        "訂單號碼",
        "提單/報關號碼",
        "幣別",
        "原幣金額",
        "匯率",
        "台幣金額",
        "結帳方式",
        "結帳期間",
        "交易日期",
        "到期日",
        "已收/已付金額",
        "未結金額",
        "逾期天數",
        "承辦人",
        "備註",
    ]
    view = df[[col for col in visible_cols if col in df.columns]].copy()
    st.dataframe(format_money_columns(view), hide_index=True, use_container_width=True, height=520)

    st.subheader("客戶/供應商彙總")
    summary = (
        df.pivot_table(
            index=["客戶/供應商", "應收/應付"],
            values=["台幣金額", "已收/已付金額", "未結金額"],
            aggfunc="sum",
            fill_value=0,
        )
        .reset_index()
        .sort_values("未結金額", ascending=False)
    )
    st.dataframe(format_money_columns(summary), hide_index=True, use_container_width=True)


def render_excel_tools(raw_df: pd.DataFrame) -> None:
    col1, col2 = st.columns(2)
    with col1:
        st.download_button(
            "下載 Excel 匯入範本",
            data=build_template_workbook(),
            file_name=f"{COMPANY_SHORT_NAME}_匯入範本.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with col2:
        report_df = enrich_transactions(raw_df)
        customers_df = fetch_customers()
        st.download_button(
            "下載完整對帳 Excel",
            data=build_report_workbook(report_df),
            file_name=report_file_name(),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            disabled=report_df.empty and customers_df.empty,
            on_click=mark_data_downloaded,
        )

    st.divider()
    uploaded = st.file_uploader("上傳 Excel", type=["xlsx", "xls"])
    import_mode = st.radio("匯入模式", ["追加或更新", "取代全部資料"], horizontal=True)
    if uploaded is not None:
        try:
            payload = parse_workbook_payload(uploaded)
            records = payload["transactions"]
            customers = payload["customers"]
            if records:
                st.markdown("**交易明細預覽**")
                preview = enrich_transactions(pd.DataFrame(records))
                st.dataframe(preview.head(20), hide_index=True, use_container_width=True)
            if customers:
                st.markdown("**ERP 客戶主檔預覽**")
                st.dataframe(format_customer_preview(pd.DataFrame(customers).head(20)), hide_index=True, use_container_width=True)
            if st.button("確認匯入", type="primary"):
                if records:
                    if import_mode == "取代全部資料":
                        replace_transactions(records)
                    else:
                        append_transactions(records)
                if customers:
                    replace_customers(customers)
                mark_data_changed()
                st.success(f"已匯入 {len(records)} 筆交易、{len(customers)} 筆 ERP 客戶資料。")
                st.rerun()
        except Exception as exc:
            st.error(f"匯入失敗：{exc}")


def read_upload_bytes(uploaded_file: Any) -> bytes:
    if isinstance(uploaded_file, (str, Path)):
        return Path(uploaded_file).read_bytes()
    if hasattr(uploaded_file, "getvalue"):
        return uploaded_file.getvalue()
    if hasattr(uploaded_file, "read"):
        position = uploaded_file.tell() if hasattr(uploaded_file, "tell") else None
        data = uploaded_file.read()
        if position is not None and hasattr(uploaded_file, "seek"):
            uploaded_file.seek(position)
        return data
    raise ValueError("無法讀取上傳檔案。")


def parse_workbook_payload(uploaded_file: Any) -> dict[str, list[dict[str, Any]]]:
    data = read_upload_bytes(uploaded_file)
    transactions: list[dict[str, Any]] = []
    customers: list[dict[str, Any]] = []

    try:
        transactions = parse_uploaded_workbook(io.BytesIO(data))
    except Exception:
        transactions = []

    try:
        customers = parse_customer_workbook(io.BytesIO(data))
    except Exception:
        customers = []

    if not transactions and not customers:
        raise ValueError("這份 Excel 沒有可匯入的交易明細，也不是可辨識的 ERP 客戶主檔。")
    return {"transactions": transactions, "customers": customers}


def parse_customer_workbook(uploaded_file: Any) -> list[dict[str, Any]]:
    workbook = pd.ExcelFile(uploaded_file)
    records: list[dict[str, Any]] = []
    seen: set[str] = set()

    for sheet_name in workbook.sheet_names:
        df = pd.read_excel(workbook, sheet_name=sheet_name, dtype=object)
        df = df.dropna(how="all")
        if df.empty:
            continue
        normalized_columns = {}
        for col in df.columns:
            key = clean_text(col)
            if key in CUSTOMER_ALIASES:
                normalized_columns[col] = CUSTOMER_ALIASES[key]
        normalized = df.rename(columns=normalized_columns)
        if not {"customer_id", "english_name"}.issubset(set(normalized.columns)):
            continue

        for _, row in normalized.iterrows():
            customer_id = clean_text(row.get("customer_id"))
            english_name = clean_text(row.get("english_name"))
            if not customer_id or not english_name or customer_id in seen:
                continue
            credit_days = int(safe_float(row.get("credit_days")))
            grace_days = int(safe_float(row.get("grace_days"), default=-1))
            if grace_days < 0:
                grace_days = infer_grace_days_from_terms(row.get("payment_terms"), credit_days)
            settlement_cycle = normalize_cycle(row.get("settlement_cycle") or ("當下結" if grace_days <= 0 else "月結"))
            records.append(
                {
                    "customer_id": customer_id,
                    "english_name": english_name,
                    "chinese_name": clean_text(row.get("chinese_name")),
                    "currency": normalize_currency(row.get("currency")),
                    "credit_days": credit_days,
                    "settlement_cycle": settlement_cycle,
                    "grace_days": grace_days,
                    "payment_terms": clean_text(row.get("payment_terms")),
                    "sales_person": clean_text(row.get("sales_person")),
                    "business_type": clean_text(row.get("business_type")),
                    "shipment_terms": clean_text(row.get("shipment_terms")),
                    "contact_person": clean_text(row.get("contact_person")),
                    "phone": clean_text(row.get("phone")),
                    "email": clean_text(row.get("email")),
                }
            )
            seen.add(customer_id)

    return records


def parse_uploaded_workbook(uploaded_file: Any) -> list[dict[str, Any]]:
    workbook = pd.ExcelFile(uploaded_file)
    sheet_name = "交易明細" if "交易明細" in workbook.sheet_names else workbook.sheet_names[0]
    df = pd.read_excel(workbook, sheet_name=sheet_name)
    df = df.dropna(how="all")
    if df.empty:
        raise ValueError("Excel 沒有可匯入的資料。")

    normalized_columns = {}
    for col in df.columns:
        key = clean_text(col)
        if key in IMPORT_ALIASES:
            normalized_columns[col] = IMPORT_ALIASES[key]
    df = df.rename(columns=normalized_columns)

    missing = [col for col in ["counterparty", "amount_original", "invoice_date"] if col not in df.columns]
    if missing:
        labels = ", ".join(DISPLAY_COLUMNS[col] for col in missing)
        raise ValueError(f"缺少必要欄位：{labels}")

    records: list[dict[str, Any]] = []
    for _, row in df.iterrows():
        if row.dropna().empty:
            continue
        trade_flow = normalize_choice(row.get("trade_flow"), TRADE_FLOWS, "出口")
        default_side = "應收" if trade_flow == "出口" else "應付"
        invoice_date = coerce_date(row.get("invoice_date")) or today()
        settlement_cycle = normalize_cycle(row.get("settlement_cycle"))
        grace_days = int(safe_float(row.get("grace_days")))
        exchange_rate = safe_float(row.get("exchange_rate"), default=0.0)
        if exchange_rate <= 0:
            exchange_rate, _ = get_live_exchange_rate(clean_text(row.get("currency")).upper() or "TWD")
        amount_original = safe_float(row.get("amount_original"))
        amount_twd = safe_float(row.get("amount_twd"), default=amount_original * exchange_rate)
        if amount_original <= 0 and amount_twd <= 0:
            continue
        if amount_original <= 0 and exchange_rate > 0:
            amount_original = round(amount_twd / exchange_rate, 2)

        records.append(
            {
                "id": clean_text(row.get("id")) or str(uuid.uuid4()),
                "trade_flow": trade_flow,
                "account_side": normalize_choice(row.get("account_side"), ACCOUNT_SIDES, default_side),
                "customer_id": clean_text(row.get("customer_id")),
                "counterparty": clean_text(row.get("counterparty")) or "未命名",
                "invoice_no": clean_text(row.get("invoice_no")),
                "order_no": clean_text(row.get("order_no")),
                "shipment_no": clean_text(row.get("shipment_no")),
                "item_description": clean_text(row.get("item_description")),
                "currency": normalize_currency(row.get("currency")),
                "amount_original": amount_original,
                "exchange_rate": exchange_rate,
                "settlement_cycle": settlement_cycle,
                "invoice_date": invoice_date,
                "grace_days": grace_days,
                "paid_amount_twd": safe_float(row.get("paid_amount_twd")),
                "payment_date": coerce_date(row.get("payment_date")),
                "bank_account": clean_text(row.get("bank_account")),
                "owner": clean_text(row.get("owner")),
                "notes": clean_text(row.get("notes")),
            }
        )
    if not records:
        raise ValueError("Excel 沒有可匯入的有效金額資料。")
    return records


def build_template_workbook() -> bytes:
    columns = [
        "進出口",
        "應收/應付",
        "客戶編號",
        "客戶/供應商",
        "發票號碼",
        "訂單號碼",
        "提單/報關號碼",
        "品名/摘要",
        "幣別",
        "原幣金額",
        "匯率",
        "結帳方式",
        "交易日期",
        "付款天數",
        "已收/已付金額",
        "收付款日期",
        "銀行/帳戶",
        "承辦人",
        "備註",
    ]
    template = pd.DataFrame(columns=columns)
    field_notes = pd.DataFrame(
        [
            ["進出口", "出口或進口"],
            ["應收/應付", "出口通常是應收，進口通常是應付，可依實際情況調整"],
            ["客戶編號", "可填 ERP Customer ID；若先匯入 ERP 客戶主檔，新增帳款時可自動帶入"],
            ["結帳方式", "當下結、月結、雙月結、半年結"],
            ["付款天數", "結帳日後再加幾天付款，例如月結 30 天填 30"],
            ["匯率", "台幣金額會以原幣金額乘以匯率計算"],
            ["ID", "匯出後若保留 ID，再匯入會更新同一筆資料"],
        ],
        columns=["欄位", "說明"],
    )
    customer_template = pd.DataFrame(
        columns=[
            "客戶編號",
            "客戶英文名稱",
            "中文名稱",
            "幣別",
            "ERP信用天數",
            "結帳方式",
            "付款天數",
            "付款條件",
            "業務",
            "客戶類別",
            "出貨條件",
            "聯絡人",
            "電話",
            "Email",
        ]
    )
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        template.to_excel(writer, sheet_name="交易明細", index=False)
        customer_template.to_excel(writer, sheet_name="ERP客戶主檔", index=False)
        field_notes.to_excel(writer, sheet_name="欄位說明", index=False)
        apply_workbook_style(writer.book)
    return output.getvalue()


def customer_export_frame(customers: pd.DataFrame | None = None) -> pd.DataFrame:
    customers_df = fetch_customers() if customers is None else customers.copy()
    export_columns = {
        "customer_id": "客戶編號",
        "english_name": "客戶英文名稱",
        "chinese_name": "中文名稱",
        "currency": "幣別",
        "credit_days": "ERP信用天數",
        "settlement_cycle": "結帳方式",
        "grace_days": "付款天數",
        "payment_terms": "付款條件",
        "sales_person": "業務",
        "business_type": "客戶類別",
        "shipment_terms": "出貨條件",
        "contact_person": "聯絡人",
        "phone": "電話",
        "email": "Email",
        "imported_at": "匯入時間",
    }
    ordered = list(export_columns.values())
    if customers_df.empty:
        return pd.DataFrame(columns=ordered)

    for col in export_columns:
        if col not in customers_df.columns:
            customers_df[col] = ""
    customers_df["currency"] = customers_df["currency"].apply(normalize_currency)
    customers_df["settlement_cycle"] = customers_df["settlement_cycle"].apply(normalize_cycle)
    customers_df["credit_days"] = pd.to_numeric(customers_df["credit_days"], errors="coerce").fillna(0).astype(int)
    customers_df["grace_days"] = pd.to_numeric(customers_df["grace_days"], errors="coerce").fillna(0).astype(int)
    renamed = customers_df.rename(columns=export_columns)
    return renamed[ordered]


def build_report_workbook(df: pd.DataFrame) -> bytes:
    output = io.BytesIO()
    customers = customer_export_frame()
    if df.empty:
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            pd.DataFrame(
                [
                    ["公司名稱", COMPANY_NAME, ""],
                    ["報表日期", today().isoformat(), ""],
                    ["交易筆數", 0, "目前沒有交易明細"],
                    ["ERP客戶筆數", len(customers), "已載入的客戶主檔會保留在 ERP客戶主檔 工作表"],
                ],
                columns=["項目", "數值", "說明"],
            ).to_excel(writer, sheet_name="對帳總表", index=False)
            pd.DataFrame(columns=list(DISPLAY_COLUMNS.values())).to_excel(writer, sheet_name="交易明細", index=False)
            customers.to_excel(writer, sheet_name="ERP客戶主檔", index=False)
            apply_workbook_style(writer.book)
        return output.getvalue()

    report = make_excel_safe(df.copy())
    for text_col in ["客戶編號", "客戶/供應商"]:
        if text_col in report.columns:
            report[text_col] = report[text_col].fillna("")
    summary = build_reconciliation_summary(report)
    aging = (
        report.pivot_table(
            index=["應收/應付", "帳齡"],
            values="未結金額",
            aggfunc="sum",
            fill_value=0,
        )
        .reset_index()
        .sort_values(["應收/應付", "帳齡"])
    )
    cycle = (
        report.pivot_table(
            index=["結帳方式", "結帳期間", "應收/應付"],
            values=["台幣金額", "已收/已付金額", "未結金額"],
            aggfunc="sum",
            fill_value=0,
        )
        .reset_index()
        .sort_values(["結帳方式", "結帳期間", "應收/應付"])
    )
    counterparty = (
        report.pivot_table(
            index=["客戶編號", "客戶/供應商", "應收/應付"],
            values=["台幣金額", "已收/已付金額", "未結金額"],
            aggfunc="sum",
            fill_value=0,
        )
        .reset_index()
        .sort_values("未結金額", ascending=False)
    )

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="對帳總表", index=False)
        report.to_excel(writer, sheet_name="交易明細", index=False)
        customers.to_excel(writer, sheet_name="ERP客戶主檔", index=False)
        aging.to_excel(writer, sheet_name="帳齡分析", index=False)
        cycle.to_excel(writer, sheet_name="結帳週期", index=False)
        counterparty.to_excel(writer, sheet_name="客戶供應商", index=False)
        apply_workbook_style(writer.book)
    return output.getvalue()


def make_excel_safe(df: pd.DataFrame) -> pd.DataFrame:
    safe = df.copy()
    for col in safe.columns:
        if pd.api.types.is_datetime64_any_dtype(safe[col]):
            if getattr(safe[col].dt, "tz", None) is not None:
                safe[col] = safe[col].dt.tz_localize(None)
        elif safe[col].dtype == "object":
            safe[col] = safe[col].apply(strip_timezone)
    return safe


def strip_timezone(value: Any) -> Any:
    if isinstance(value, pd.Timestamp):
        return value.tz_localize(None) if value.tz is not None else value
    if isinstance(value, datetime) and value.tzinfo is not None:
        return value.replace(tzinfo=None)
    return value


def build_reconciliation_summary(df: pd.DataFrame) -> pd.DataFrame:
    receivable = df[df["應收/應付"] == "應收"]["未結金額"].sum()
    payable = df[df["應收/應付"] == "應付"]["未結金額"].sum()
    overdue = df[df["逾期天數"] > 0]["未結金額"].sum()
    settled = df[df["狀態"] == "已結清"]["台幣金額"].sum()
    return pd.DataFrame(
        [
            ["公司名稱", COMPANY_NAME, ""],
            ["統一編號", COMPANY_TAX_ID, ""],
            ["聯絡電話", COMPANY_PHONE, ""],
            ["報表日期", today().isoformat(), ""],
            ["交易筆數", len(df), ""],
            ["應收未結", receivable, "應收帳款尚未收款"],
            ["應付未結", payable, "應付帳款尚未付款"],
            ["淨部位", receivable - payable, "應收未結減應付未結"],
            ["逾期未結", overdue, "超過到期日仍未結清"],
            ["已結清金額", settled, "已完全收/付完成的台幣金額"],
        ],
        columns=["項目", "數值", "說明"],
    )


def apply_workbook_style(workbook: Any) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="284B46")
    header_font = Font(color="FFFFFF", bold=True)
    money_headers = {"原幣金額", "台幣金額", "已收/已付金額", "未結金額", "數值"}

    for ws in workbook.worksheets:
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False
        if ws.max_row >= 1 and ws.max_column >= 1:
            ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for col_idx, column_cells in enumerate(ws.columns, start=1):
            header = column_cells[0].value
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 34)
            if header in money_headers:
                for cell in column_cells[1:]:
                    cell.number_format = '#,##0.00'
            if "日期" in str(header) or header in {"交易日期", "到期日", "收付款日期", "建立時間", "更新時間", "匯入時間"}:
                for cell in column_cells[1:]:
                    cell.number_format = "yyyy-mm-dd"


def render_settings() -> None:
    st.subheader("公司資訊")
    company_info = pd.DataFrame(
        [
            ["公司名稱", COMPANY_NAME],
            ["統一編號", COMPANY_TAX_ID],
            ["聯絡電話", COMPANY_PHONE],
            ["公司地址", COMPANY_ADDRESS],
            ["公司網站", COMPANY_WEBSITE],
            ["營業項目", "、".join(COMPANY_BUSINESS_LINES)],
        ],
        columns=["項目", "內容"],
    )
    st.dataframe(company_info, hide_index=True, use_container_width=True)

    st.subheader("結帳規則")
    rules = pd.DataFrame(
        [
            ["當下結", "交易日當天", "交易日 + 付款天數"],
            ["月結", "交易月份最後一天", "月底 + 付款天數"],
            ["雙月結", "1-2、3-4、5-6、7-8、9-10、11-12 月分組", "該雙月最後一天 + 付款天數"],
            ["半年結", "1-6 月、7-12 月分組", "6/30 或 12/31 + 付款天數"],
        ],
        columns=["結帳方式", "結帳日", "到期日算法"],
    )
    st.dataframe(rules, hide_index=True, use_container_width=True)

    st.subheader("本機資料庫")
    db_path = Path(os.getenv("ACCOUNTING_DB_PATH", str(DATA_DIR / "accounting.db")))
    st.code(str(db_path), language="text")

    st.subheader("示範資料")
    if st.button("載入示範資料", type="secondary"):
        append_transactions(sample_records())
        mark_data_changed()
        st.success("已載入示範資料。")
        st.rerun()


def sample_records() -> list[dict[str, Any]]:
    base = today()
    return [
        {
            "trade_flow": "出口",
            "account_side": "應收",
            "counterparty": "Tokyo Trading Co.",
            "invoice_no": "EX-2026-001",
            "order_no": "SO-26001",
            "shipment_no": "BL-TYO-001",
            "item_description": "電子零件",
            "currency": "USD",
            "amount_original": 18500,
            "exchange_rate": 32.15,
            "settlement_cycle": "月結",
            "invoice_date": base.replace(day=5),
            "grace_days": 30,
            "paid_amount_twd": 0,
            "owner": "業務部",
            "notes": "樣本資料",
        },
        {
            "trade_flow": "進口",
            "account_side": "應付",
            "counterparty": "Shenzhen Supply Ltd.",
            "invoice_no": "IM-2026-019",
            "order_no": "PO-26019",
            "shipment_no": "DECL-26019",
            "item_description": "包材",
            "currency": "CNY",
            "amount_original": 72000,
            "exchange_rate": 4.42,
            "settlement_cycle": "雙月結",
            "invoice_date": base - timedelta(days=46),
            "grace_days": 15,
            "paid_amount_twd": 120000,
            "owner": "採購部",
            "notes": "樣本資料",
        },
        {
            "trade_flow": "出口",
            "account_side": "應收",
            "counterparty": "Pacific Buyer Inc.",
            "invoice_no": "EX-2026-002",
            "order_no": "SO-26002",
            "shipment_no": "BL-LAX-002",
            "item_description": "機械配件",
            "currency": "USD",
            "amount_original": 42000,
            "exchange_rate": 32.1,
            "settlement_cycle": "半年結",
            "invoice_date": base - timedelta(days=85),
            "grace_days": 30,
            "paid_amount_twd": 0,
            "owner": "業務部",
            "notes": "樣本資料",
        },
    ]


def format_money_columns(df: pd.DataFrame) -> pd.DataFrame:
    formatted = df.copy()
    for col in formatted.columns:
        if col in {"原幣金額", "台幣金額", "已收/已付金額", "未結金額", "應收", "應付"}:
            formatted[col] = formatted[col].apply(money)
    return formatted


def format_customer_preview(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    renamed = df.rename(
        columns={
            "customer_id": "客戶編號",
            "english_name": "客戶英文名稱",
            "currency": "幣別",
            "credit_days": "ERP信用天數",
            "settlement_cycle": "結帳方式",
            "grace_days": "付款天數",
            "payment_terms": "付款條件",
            "sales_person": "業務",
            "business_type": "客戶類別",
        }
    )
    cols = ["客戶編號", "客戶英文名稱", "幣別", "ERP信用天數", "結帳方式", "付款天數", "付款條件", "業務", "客戶類別"]
    return renamed[[col for col in cols if col in renamed.columns]]


def money(value: Any) -> str:
    return f"{safe_float(value):,.0f}"


def today() -> date:
    return datetime.now(APP_TZ).date()


def clean_text(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    return str(value).strip()


def safe_float(value: Any, default: float = 0.0) -> float:
    if value is None or value == "":
        return default
    try:
        if pd.isna(value):
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def coerce_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.date()


def normalize_choice(value: Any, choices: list[str], default: str) -> str:
    cleaned = clean_text(value)
    return cleaned if cleaned in choices else default


def normalize_cycle(value: Any) -> str:
    cleaned = clean_text(value)
    aliases = {
        "現結": "當下結",
        "即期": "當下結",
        "當日結": "當下結",
        "每月結": "月結",
        "二月結": "雙月結",
        "雙月份結": "雙月結",
        "半年度結": "半年結",
    }
    cleaned = aliases.get(cleaned, cleaned)
    return cleaned if cleaned in SETTLEMENT_CYCLES else "月結"


def normalize_currency(value: Any) -> str:
    cleaned = clean_text(value).upper()
    aliases = {
        "NTD": "TWD",
        "NT": "TWD",
        "RMB": "CNY",
        "人民幣": "CNY",
        "台幣": "TWD",
        "新台幣": "TWD",
    }
    normalized = aliases.get(cleaned, cleaned)
    return normalized if normalized in CURRENCIES else "TWD"


def infer_grace_days_from_terms(payment_terms: Any, credit_days: Any) -> int:
    text = clean_text(payment_terms)
    numbers = [int(match) for match in re.findall(r"\d+", text)]
    if numbers:
        return max(numbers)
    return max(int(safe_float(credit_days)), 0)


def index_of(options: list[str], value: Any, fallback: int) -> int:
    cleaned = clean_text(value)
    return options.index(cleaned) if cleaned in options else fallback


if __name__ == "__main__":
    main()
